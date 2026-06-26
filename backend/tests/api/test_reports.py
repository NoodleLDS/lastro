from io import BytesIO

from httpx import AsyncClient
from openpyxl import load_workbook


async def test_report_default_modules_retorna_json_com_todas_as_chaves(
    client: AsyncClient,
) -> None:
    response = await client.get("/reports")
    assert response.status_code == 200
    body = response.json()
    for key in [
        "transactions",
        "contributions",
        "dividends",
        "sales",
        "incomes",
        "fixed_expenses",
        "variable_expenses",
        "positions",
    ]:
        assert key in body


async def test_report_filtra_por_modulo(client: AsyncClient) -> None:
    response = await client.get("/reports", params={"modules": ["transactions"]})
    assert response.status_code == 200
    body = response.json()
    assert body["transactions"] == []
    assert body["contributions"] is None
    assert body["positions"] is None


async def test_report_inclui_transacao_dentro_do_periodo(client: AsyncClient) -> None:
    card = (await client.post("/cards", json={"name": "Nubank"})).json()
    await client.post(
        "/transactions/quick-entry",
        json={"card_id": card["id"], "raw": "mercado 100", "date": "2026-06-10"},
    )

    response = await client.get(
        "/reports",
        params={
            "modules": ["transactions"],
            "date_from": "2026-06-01",
            "date_to": "2026-06-30",
        },
    )
    body = response.json()
    assert len(body["transactions"]) == 1
    assert body["transactions"][0]["description"] == "mercado"

    response = await client.get(
        "/reports",
        params={"modules": ["transactions"], "date_from": "2026-07-01", "date_to": "2026-07-31"},
    )
    assert response.json()["transactions"] == []


async def test_report_filtra_incomes_por_periodo(client: AsyncClient) -> None:
    await client.post(
        "/incomes",
        json={"description": "Salário", "amount_cents": 500_000, "year": 2026, "month": 6},
    )

    response = await client.get(
        "/reports",
        params={
            "modules": ["incomes"],
            "date_from": "2026-06-01",
            "date_to": "2026-06-30",
        },
    )
    assert len(response.json()["incomes"]) == 1

    response = await client.get(
        "/reports",
        params={"modules": ["incomes"], "date_from": "2026-01-01", "date_to": "2026-01-31"},
    )
    assert response.json()["incomes"] == []


async def test_report_xlsx_retorna_planilha_valida(client: AsyncClient) -> None:
    card = (await client.post("/cards", json={"name": "Inter"})).json()
    await client.post(
        "/transactions/quick-entry",
        json={"card_id": card["id"], "raw": "posto 50", "date": "2026-06-15"},
    )

    response = await client.get(
        "/reports",
        params={"modules": ["transactions"], "format": "xlsx"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["transactions"]
    sheet = workbook["transactions"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][1] == "date"
    assert any(row[2] == "posto" for row in rows[1:])
