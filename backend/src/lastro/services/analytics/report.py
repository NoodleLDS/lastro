from datetime import date as date_
from io import BytesIO

from openpyxl import Workbook
from sqlmodel import select
from sqlmodel.ext.asyncio.session import AsyncSession

from lastro.models.contribution import Contribution
from lastro.models.dividend import Dividend
from lastro.models.fixed_expense import FixedExpense
from lastro.models.income import Income
from lastro.models.position import Position
from lastro.models.sale import Sale
from lastro.models.transaction import Transaction, TransactionStatus
from lastro.models.variable_expense import VariableExpense
from lastro.schemas.contribution import ContributionRead
from lastro.schemas.dividend import DividendRead
from lastro.schemas.monthly_summary import FixedExpenseRead, IncomeRead, VariableExpenseRead
from lastro.schemas.report import PositionSnapshot, Report, ReportModule
from lastro.schemas.sale import SaleRead
from lastro.schemas.transaction import TransactionRead


def _year_month_bounds(
    date_from: date_ | None, date_to: date_ | None
) -> tuple[int | None, int | None, int | None, int | None]:
    """Converte um intervalo de datas em (year_from, month_from, year_to, month_to)."""
    year_from = date_from.year if date_from else None
    month_from = date_from.month if date_from else None
    year_to = date_to.year if date_to else None
    month_to = date_to.month if date_to else None
    return year_from, month_from, year_to, month_to


def _within_year_month(
    year: int, month: int, year_from: int | None, month_from: int | None,
    year_to: int | None, month_to: int | None,
) -> bool:
    key = (year, month)
    if year_from is not None and month_from is not None and key < (year_from, month_from):
        return False
    if year_to is not None and month_to is not None and key > (year_to, month_to):
        return False
    return True


async def build_report(
    session: AsyncSession,
    modules: list[ReportModule],
    date_from: date_ | None,
    date_to: date_ | None,
    category_id: int | None,
) -> Report:
    report = Report()
    year_from, month_from, year_to, month_to = _year_month_bounds(date_from, date_to)

    if ReportModule.TRANSACTIONS in modules:
        statement = select(Transaction).where(Transaction.status == TransactionStatus.CONFIRMED)
        if category_id is not None:
            statement = statement.where(Transaction.category_id == category_id)
        if date_from is not None:
            statement = statement.where(Transaction.date >= date_from)
        if date_to is not None:
            statement = statement.where(Transaction.date <= date_to)
        result = await session.exec(statement.order_by(Transaction.date))
        report.transactions = [TransactionRead.model_validate(t) for t in result.all()]

    if ReportModule.CONTRIBUTIONS in modules:
        statement = select(Contribution)
        if date_from is not None:
            statement = statement.where(Contribution.date >= date_from)
        if date_to is not None:
            statement = statement.where(Contribution.date <= date_to)
        result = await session.exec(statement.order_by(Contribution.date))
        report.contributions = [ContributionRead.model_validate(c) for c in result.all()]

    if ReportModule.DIVIDENDS in modules:
        statement = select(Dividend)
        if date_from is not None:
            statement = statement.where(Dividend.date >= date_from)
        if date_to is not None:
            statement = statement.where(Dividend.date <= date_to)
        result = await session.exec(statement.order_by(Dividend.date))
        report.dividends = [DividendRead.model_validate(d) for d in result.all()]

    if ReportModule.SALES in modules:
        statement = select(Sale)
        if date_from is not None:
            statement = statement.where(Sale.date >= date_from)
        if date_to is not None:
            statement = statement.where(Sale.date <= date_to)
        result = await session.exec(statement.order_by(Sale.date))
        report.sales = [SaleRead.model_validate(s) for s in result.all()]

    if ReportModule.INCOMES in modules:
        statement = select(Income)
        result = await session.exec(statement)
        incomes = [
            i for i in result.all()
            if _within_year_month(i.year, i.month, year_from, month_from, year_to, month_to)
        ]
        report.incomes = [IncomeRead.model_validate(i) for i in incomes]

    if ReportModule.FIXED_EXPENSES in modules:
        statement = select(FixedExpense)
        if category_id is not None:
            statement = statement.where(FixedExpense.category_id == category_id)
        result = await session.exec(statement)
        expenses = [
            e for e in result.all()
            if _within_year_month(e.year, e.month, year_from, month_from, year_to, month_to)
        ]
        report.fixed_expenses = [FixedExpenseRead.model_validate(e) for e in expenses]

    if ReportModule.VARIABLE_EXPENSES in modules:
        statement = select(VariableExpense)
        if category_id is not None:
            statement = statement.where(VariableExpense.category_id == category_id)
        result = await session.exec(statement)
        expenses = [
            e for e in result.all()
            if _within_year_month(e.year, e.month, year_from, month_from, year_to, month_to)
        ]
        report.variable_expenses = [VariableExpenseRead.model_validate(e) for e in expenses]

    if ReportModule.POSITIONS in modules:
        statement = select(Position).where(Position.is_active.is_(True))
        result = await session.exec(statement)
        report.positions = [
            PositionSnapshot(
                id=p.id,
                ticker=p.ticker,
                name=p.name,
                asset_type=p.asset_type.value,
                quantity=p.quantity,
                last_price_cents=p.last_price_cents,
            )
            for p in result.all()
        ]

    return report


def render_xlsx(report: Report) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for field_name in Report.model_fields:
        records = getattr(report, field_name)
        if records is None:
            continue

        sheet = workbook.create_sheet(title=field_name[:31])
        if not records:
            continue

        headers = list(records[0].model_dump().keys())
        sheet.append(headers)
        for record in records:
            row = record.model_dump()
            sheet.append([row[h] for h in headers])

    if not workbook.sheetnames:
        workbook.create_sheet(title="vazio")

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
